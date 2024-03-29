# -*- coding: utf-8 -*-
from datetime import datetime
from pathlib import Path
from typing import Protocol, Sequence, Iterable, Tuple, List, Callable

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

import config
from exceptions import FormatToStrError
from quake_structures import Quake


class QuakesStorage(Protocol):
    """Interface of any storage for saving info of quakes"""

    def save(self, quakes: Iterable[Quake]) -> None:
        raise NotImplementedError


def save_quakes(quakes: Iterable[Quake], storage: QuakesStorage) -> None:
    """Save quakes in the storage"""
    storage.save(quakes)


class CatalogStorage(QuakesStorage):
    """Store some quakes info as a catalog in an Excel file"""

    def __init__(self, file: Path):
        self._file = file
        self._init_storage()
        self.wb = openpyxl.load_workbook(self._file)
        self._del_init_empty_worksheet()
        self.sheet: Worksheet

    def save(self, quakes: Iterable[Quake]) -> None:
        for quake in quakes:
            if quake.lat is None or quake.lon is None:
                continue
            self._init_sheet(quake)
            self._add_values_in_sheet(quake)
        self._format_cells()
        self.wb.save(self._file)

    def _init_storage(self) -> None:
        if not self._file.exists():
            wb = openpyxl.Workbook()
            wb.save(self._file)

    def _add_values_in_sheet(self, quake: Quake) -> None:
        origin_dt = _format_common_attrs(quake)[0]
        origin_d, origin_t = origin_dt.split()
        lat = quake.lat if quake.lat else '-'
        lon = quake.lon if quake.lon else '-'
        mag = quake.magnitude
        avg_ml = mag.ML if mag.ML != 0.0 else '-'
        avg_mpsp = mag.MPSP if mag.MPSP != 0.0 else '-'
        depth = quake.depth if quake.depth else '-'
        stations_name = ', '.join(quake.stations_name)
        row = (origin_d, origin_t, lat, lon, depth,
               quake.reg, avg_ml, avg_mpsp, stations_name)
        self.sheet.append(row)

    def _format_cells(self) -> None:
        for sheet in self.wb.sheetnames:
            rows_of_cells = self.wb[sheet][self.wb[sheet].dimensions]
            for row in rows_of_cells:
                for cell in row:
                    if cell.column in (3, 4, 5):
                        # digital format for lat, lon, depth
                        cell.number_format = '0.00'
                    elif cell.column in (7, 8):
                        cell.number_format = '0.0'
                    cell.alignment = Alignment(horizontal='center',
                                               vertical='center')

    def _init_sheet(self, quake: Quake) -> None:
        month_num = quake.origin_dt.month - 1
        sheet_name = config.MONTHS[month_num]
        if sheet_name not in self.wb.sheetnames:
            self.sheet = self.wb.create_sheet(sheet_name, month_num)
            self.sheet.append(config.CATALOG_HEADER)
        else:
            self.sheet = self.wb.get_sheet_by_name(sheet_name)

    def _del_init_empty_worksheet(self):
        if 'Sheet' in self.wb.sheetnames and \
                self.wb['Sheet'].dimensions == 'A1:A1':
            del self.wb['Sheet']


class BulletinStorage(QuakesStorage):
    """Store some quakes info as a bulletin in a plain text file"""

    def __init__(self, file: Path):
        self._file = file
        self.origin_dt = ''
        self.lat = ''
        self.lon = ''
        self.mag = ''
        self.avg_ml = ''
        self.avg_mpsp = ''
        self.mag_type = '-'
        self.depth = ''

    def save(self, quakes: Iterable[Quake]) -> None:
        with self._file.open('w', encoding='utf8') as f:
            amnt_quakes = 0
            for quake in quakes:
                (self.origin_dt, self.lat, self.lon, self.mag, self.avg_ml,
                 self.avg_mpsp, self.depth,
                 self.mag_type) = _format_common_attrs(quake)
                rows = self._get_rows(quake)
                f.write('\n'.join(rows))
                amnt_quakes += 1
            f.write(f'\nTotal: {amnt_quakes}')

    def _get_rows(self, quake: Quake) -> Iterable[str]:
        quake_hdr_describe = self._get_quake_hdr_describe()
        quake_hdr = self._get_quake_hdr(quake) + '\n'
        sta_hdr_describe = \
            _format_to_str(config.STATION_HEADER_DESCRIBE,
                           config.AMNT_COLUMN_SYMBOLS['sta_hdr'])
        sta_strings = self._get_stations_string(quake)
        return ('#' + quake.id, quake_hdr_describe, quake_hdr,
                sta_hdr_describe, sta_strings)

    def _get_quake_hdr_describe(self) -> str:
        mag_type = 'Mag' if self.mag_type == '-' else self.mag_type
        columns_data = config.QUAKE_HEADER_DESCRIBE[:]
        columns_data.insert(5, mag_type)
        return _format_to_str(columns_data,
                              config.AMNT_COLUMN_SYMBOLS['quake_hdr'])

    def _get_quake_hdr(self, quake: Quake) -> str:
        amnt_sta = str(len(quake.stations_name))
        return _format_to_str(
            columns_data=(self.origin_dt, self.lat, self.lon, self.depth,
                          amnt_sta, self.mag, quake.reg),
            hdr_type_config=config.AMNT_COLUMN_SYMBOLS['quake_hdr'])

    def _get_stations_string(self, quake: Quake) -> str:
        res = ''
        for sta in quake.stations:
            phase_dt = datetime.strftime(sta.phase_dt,
                                         '%d.%m.%Y %H:%M:%S.%f')[:-3]
            dist = f'{sta.dist:.2f}' if sta.dist else '-'
            az = f'{sta.azimuth:.2f}' if sta.azimuth else '-'
            ampl = f'{sta.ampl:.4f}' if sta.ampl else '-'
            period = f'{sta.period:.2f}' if sta.period else '-'
            mag = f'{sta.mag_ML:.1f}' if sta.mag_ML else \
                f'{sta.mag_MPSP:.1f}' if sta.mag_MPSP else '-'
            mag_type = 'ML' if sta.mag_ML else 'MPSP' if sta.mag_MPSP else '-'
            sta_data = (sta.name, dist, az, sta.phase, sta.entry, phase_dt,
                        ampl, period, mag, mag_type)
            res += _format_to_str(sta_data,
                                  config.AMNT_COLUMN_SYMBOLS['sta_hdr']) + '\n'
        return res + '\n'


class NASBulletinStorage(QuakesStorage):
    """Store some info of each quake as a bulletin for NAS program
    in a separate plain text file with ext (*.bltn)"""

    def __init__(self, path: Path):
        self._path = path.joinpath(*path.parts[:-1])
        self.bltn_strings: List[str, ] = []

    def save(self, quakes: Iterable[Quake]) -> None:
        for quake in quakes:
            self._get_rows(quake)
            if self.bltn_strings:
                f_name = datetime.strftime(quake.origin_dt, '%Y%m%d_%H%M%S')
                full_path = self._path.joinpath(f_name).with_suffix('.bltn')
                data = '\n'.join(self.bltn_strings)
                full_path.write_text(data, encoding='utf8')

    def _get_rows(self, quake: Quake) -> None:
        self.bltn_strings.clear()
        if (quake.lat is not None and quake.lon is not None) \
                or len(quake.stations_name) > 4:
            dt, lat, lon = _format_common_attrs(quake,
                                                '%Y %m %d %H %M %S.%f')[:3]
            self.bltn_strings.append(f'Fi={lat}  LD={lon} T0={dt}')
            for sta in quake.stations:
                phase_dt = datetime.strftime(sta.phase_dt,
                                             '%Y %m %d   %H %M %S.%f')[:-3]
                self.bltn_strings.append(
                    f'{sta.name}    {sta.phase}={phase_dt}')


class ArcGisStorage(QuakesStorage):
    """Store some quakes info in a plain text file
    for further processing in ArcGis program"""

    def __init__(self, file: Path):
        self._file = file

    def save(self, quakes: Iterable[Quake]) -> None:
        with self._file.open('w', encoding='utf8') as f:
            f.write(' '.join(config.ArcGIS_HEADER) + '\n')
            for quake in quakes:
                if columns := self._get_column_values(quake):
                    row = ' '.join(columns)
                    f.write(row + '\n')

    def _get_column_values(self, quake: Quake) -> Iterable[str]:
        origin_dt, lat, lon, mag = _format_common_attrs(quake)[:4]
        if lat == '-' or lon == '-':
            return ()
        columns = origin_dt, lat, lon, '0.0', '1'
        if mag != '-':
            for _range in config.MAGNITUDE_RANGES:
                if _range[0] < float(mag) < _range[1]:
                    columns = origin_dt, lat, lon, mag, \
                              config.MAGNITUDE_RANGES[_range]
        return columns


def _format_common_attrs(quake: Quake,
                         date_fmt='%d.%m.%Y %H:%M:%S.%f') -> Tuple[str, ...]:
    origin_dt = datetime.strftime(quake.origin_dt, date_fmt)[:-3] \
        if quake.origin_dt != datetime.min else '-'
    lat = f'{quake.lat:.2f}' if quake.lat else '-'
    lon = f'{quake.lon:.2f}' if quake.lon else '-'
    mag = quake.magnitude
    avg_ml = f'{mag.ML:.1f}' if mag.ML else '-'
    avg_mpsp = f'{mag.MPSP:.1f}' if mag.MPSP else '-'
    preferred_mag = avg_ml if avg_ml != '-' else avg_mpsp
    depth = f'{quake.depth:.2f}' if quake.depth else '-'
    mag_type = 'ML' if avg_ml != '-' else 'MPSP' if avg_mpsp != '-' else '-'
    return origin_dt, lat, lon, preferred_mag, avg_ml, avg_mpsp, depth, mag_type


def _format_to_str(columns_data: Sequence, hdr_type_config: Sequence) -> str:
    """Return formatted string accordingly layout of hdr_type_config"""
    if len(columns_data) != len(hdr_type_config):
        raise FormatToStrError('len(columns_data) is not '
                               'equal len(hdr_type_config)')
    res = ''
    for i in range(len(hdr_type_config)):
        data = columns_data[i] if columns_data[i] is not None else '-'
        res += f'{data:<{hdr_type_config[i]}}'
    return res


def get_storage(ext: str) -> Callable:
    storage = {'.txt': BulletinStorage,
               '.bltn': NASBulletinStorage,
               '.xlsx': CatalogStorage,
               '.GIS': ArcGisStorage}
    return storage[ext]
