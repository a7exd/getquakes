# -*- coding: utf-8 -*-
import decimal
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import NamedTuple, Protocol, Iterable, Sequence

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

import config
from quake_handlers import format_common_attrs, format_to_str


@dataclass(slots=True)
class Sta:
    name: str
    dist: float
    azimuth: float
    phase: str
    entry: str
    phase_dt: datetime
    ampl: decimal
    period: float
    mag_ML: float
    mag_MPSP: float


class Magnitude(NamedTuple):
    """Contain average values of magnitude: ML, MPSP"""
    ML: float
    MPSP: float


@dataclass(slots=True)
class Quake:
    id: str
    origin_dt: datetime
    lat: float
    lon: float
    depth: float
    reg: str
    stations: list[Sta]

    def get_stations_name(self) -> set[Sta.name]:
        stations_name = set()
        for sta in self.stations:
            stations_name.add(sta.name)
        return stations_name

    def get_magnitude(self) -> Magnitude:
        ml = n_ml = mpsp = n_mpsp = 0
        avg_ml = avg_mpsp = None
        for sta in self.stations:
            if sta.mag_ML:
                ml += sta.mag_ML
                n_ml += 1
            if sta.mag_MPSP:
                mpsp += sta.mag_MPSP
                n_mpsp += 1
        if n_ml != 0:
            avg_ml = round(ml / n_ml, 1)
        if n_mpsp != 0:
            avg_mpsp = round(mpsp / n_mpsp, 1)
        return Magnitude(avg_ml, avg_mpsp)


class QuakesStorage(Protocol):
    """Interface of any storage for saving info of quakes"""

    def save(self, quakes: Sequence[Quake]) -> None:
        raise NotImplementedError


class CatalogStorage:
    """Store some info of quakes in Excel file"""

    def __init__(self, file: Path):
        self._file = file
        self._init_storage()
        self.wb: Workbook
        self.sheet: Worksheet

    def save(self, quakes: Sequence[Quake]) -> None:
        self._write(quakes)

    def _init_storage(self) -> None:
        if not self._file.exists():
            wb = openpyxl.Workbook()
            self._file += '.xlsx'
            wb.save(self._file)

    def _write(self, quakes: Sequence[Quake], amnt_processed: int = 0) -> None:
        month_num = quakes[amnt_processed].origin_dt.month - 1
        self._init_workbook_sheet(month_num)
        for quake in quakes:
            if quake.origin_dt.month - 1 > month_num:
                self._write(quakes[amnt_processed:], amnt_processed)
            self._add_vulues_in_sheet(quake)
            amnt_processed += 1
        self.wb.save(self._file)

    def _add_vulues_in_sheet(self, quake) -> None:
        origin_dt, lat, lon, avg_ml, avg_mpsp, depth = format_common_attrs(
            quake)
        origin_d, origin_t = origin_dt.split()
        stations_name = ', '.join(quake.get_stations_name())
        row = (origin_d, origin_t, lat, lon, depth, quake.reg, avg_ml,
               avg_mpsp, stations_name)
        if lat != '-' and lon != '-':
            self.sheet.append(row)
        columns = self.sheet.column_dimensions['A':'I']
        columns.alignment = Alignment(horizontal='center',
                                      vertical='center')

    def _init_workbook_sheet(self, month_num) -> None:
        self.wb = openpyxl.load_workbook(self._file)
        self.sheet = self.wb.create_sheet(config.MONTHS[month_num], month_num)
        self.sheet.append(config.CATALOG_HEADER)


class BulletinStorage:
    """Store some info of quakes in plain text file"""

    def __init__(self, file: Path):
        self._file = file

    def save(self, quakes: Sequence[Quake]) -> None:
        with self._file.open('w', encoding='utf8') as f:
            amnt_quakes = 0
            for quake in quakes:
                rows = self._get_rows(quake)
                f.write('\n'.join(rows))
                amnt_quakes += 1
            f.write(f'\nTotal: {amnt_quakes}')

    def _get_rows(self, quake: Quake) -> Iterable[str]:
        quake_hdr_describe = self._get_quake_hdr_describe(quake)
        quake_hdr = self._get_quake_hdr(quake) + '\n'
        sta_hdr_describe = \
            format_to_str(config.STATION_HEADER_DESCRIBE,
                          config.AMNT_COLUMN_SYMBOLS['sta_hdr'])
        sta_strings = self._get_stations_string(quake)
        return quake.id, quake_hdr_describe, quake_hdr, \
               sta_hdr_describe, sta_strings

    def _get_quake_hdr_describe(self, quake: Quake) -> str:
        mag = quake.get_magnitude()
        mag_type = 'ML' if mag.ML else 'MPSP' if mag.MPSP else 'Mag'
        config.QUAKE_HEADER_DESCRIBE.insert(5, mag_type)
        return format_to_str(config.QUAKE_HEADER_DESCRIBE,
                             config.AMNT_COLUMN_SYMBOLS['quake_hdr'])

    def _get_quake_hdr(self, quake: Quake) -> str:
        origin_dt, lat, lon, avg_ml, avg_mpsp, depth = format_common_attrs(
            quake)
        mag = avg_ml if avg_ml != '-' else avg_mpsp
        amnt_sta = str(len(quake.get_stations_name()))
        return format_to_str(
            columns_data=(origin_dt, lat, lon, depth, amnt_sta, mag, quake.reg),
            hdr_type_config=config.AMNT_COLUMN_SYMBOLS['quake_hdr'])

    def _get_stations_string(self, quake: Quake) -> str:
        res = ''
        for sta in quake.stations:
            mag = sta.mag_ML if sta.mag_ML else \
                sta.mag_MPSP if sta.mag_MPSP else '-'
            mag_type = 'ML' if sta.mag_ML else 'MPSP' if sta.mag_MPSP else '-'
            sta_data = (sta.name, sta.dist, sta.azimuth, sta.phase, sta.entry,
                        sta.phase_dt, sta.ampl, sta.period, mag, mag_type)
            res += format_to_str(sta_data,
                                 config.AMNT_COLUMN_SYMBOLS['sta_hdr']) + '\n'
        return res + '\n'


class NASBulletinStorage:
    """Store some info of each quake in its own plain text file
    with ext (*.bltn)"""

    def __init__(self, path: Path):
        self._path = path

    def save(self, quakes: Sequence[Quake]) -> None:
        for quake in quakes:
            rows = self._get_rows(quake)
            f_name = datetime.strftime(quake.origin_dt, '%Y%m%d_%H%M%S')
            full_path = Path.joinpath(self._path, f_name, '.bltn')
            with full_path.open('w', encoding='utf8') as f:
                f.write('\n'.join(rows))

    def _get_rows(self, quake: Quake) -> list[str]:
        dt = datetime.strftime(quake.origin_dt, '%Y %m %d %H %M %S.%f')[:-3]
        header = f'Fi={quake.lat:.2f}  LD={quake.lon:.2f} T0={dt}'
        bltn_strings = [header]
        for sta in quake.stations:
            phase_dt = datetime.strftime(sta.phase_dt,
                                         '%Y %m %d   %H %M %S.%f')[:-3]
            bltn_strings.append(f'{sta.name}    {sta.phase}={phase_dt}')
        return bltn_strings


class ArcGisStorage:
    """Store some info of quakes in plain text file
    for further processing in ArcGis program"""

    def __init__(self, file: Path):
        self._file = file

    def save(self, quakes: Sequence[Quake]) -> None:
        with self._file.open('w', encoding='utf8') as f:
            f.write(' '.join(config.ArcGIS_HEADER) + '\n')
            for quake in quakes:
                row = ' '.join(self._get_column_values(quake))
                f.write(row + '\n')

    def _get_column_values(self, quake: Quake) -> Iterable[str]:
        origin_dt, lat, lon, avg_ml, avg_mpsp, _ = format_common_attrs(quake)
        mag = avg_ml if avg_ml != '-' else avg_mpsp
        columns = origin_dt, lat, lon, '0.0', '1'
        if mag != '-':
            for _range in config.MAGNITUDE_RANGES:
                if _range[0] < float(mag) < _range[1]:
                    columns = origin_dt, lat, lon, mag, \
                              config.MAGNITUDE_RANGES[_range]
        return columns
